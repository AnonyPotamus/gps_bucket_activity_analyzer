import argparse
import datetime
import logging
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from google.cloud import storage
from google.oauth2.credentials import Credentials
from google.auth.exceptions import DefaultCredentialsError
import warnings
from openpyxl import Workbook
from openpyxl.styles import Font

# Suppress all warnings
warnings.filterwarnings("ignore")

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def parse_arguments():
    parser = argparse.ArgumentParser(description="Find GCS buckets not modified since a given date.")
    parser.add_argument("project_id", help="GCP Project ID")
    parser.add_argument("cutoff_date", help="Cutoff date in YYYY-MM-DD format")
    parser.add_argument("--output", default="unmodified_buckets.xlsx", help="Output Excel file name")
    parser.add_argument("--max_workers", type=int, default=10, help="Maximum number of worker threads")
    return parser.parse_args()

def get_storage_client(project_id):
    os.environ["GOOGLE_CLOUD_PROJECT"] = project_id
    try:
        credentials = Credentials.from_authorized_user_file(
            os.path.expanduser("~/.config/gcloud/application_default_credentials.json")
        )
        return storage.Client(project=project_id, credentials=credentials)
    except (DefaultCredentialsError, FileNotFoundError):
        logging.error("Unable to find local credentials. Please run 'gcloud auth application-default login'")
        raise

def check_bucket_modification(client, bucket_name, cutoff_date):
    try:
        bucket = client.get_bucket(bucket_name)
        
        if bucket.time_created.replace(tzinfo=None) > cutoff_date:
            return None  # Bucket was created after the cutoff date

        if bucket.updated and bucket.updated.replace(tzinfo=None) > cutoff_date:
            return None  # Bucket metadata was modified after the cutoff date

        # Check objects in the bucket
        blobs = client.list_blobs(bucket_name)
        for blob in blobs:
            if blob.updated.replace(tzinfo=None) > cutoff_date:
                return None  # An object was modified after the cutoff date

        return {
            'name': bucket.name,
            'created': bucket.time_created,
            'last_modified': bucket.updated or bucket.time_created,
            'location': bucket.location,
            'storage_class': bucket.storage_class
        }
    except Exception as e:
        logging.error(f"Error processing bucket {bucket_name}: {str(e)}")
        return None

def get_unmodified_buckets(client, cutoff_date, max_workers):
    unmodified_buckets = []

    try:
        buckets = list(client.list_buckets())
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_bucket = {executor.submit(check_bucket_modification, client, bucket.name, cutoff_date): bucket 
                                for bucket in buckets}
            
            for future in as_completed(future_to_bucket):
                result = future.result()
                if result:
                    unmodified_buckets.append(result)
    except Exception as e:
        logging.error(f"Error in get_unmodified_buckets: {str(e)}")

    return unmodified_buckets

def write_to_excel(buckets, filename):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Unmodified Buckets"

        # Write headers
        headers = ['Bucket Name', 'Created Date', 'Last Modified Date', 'Location', 'Storage Class']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        # Write data
        for row, bucket in enumerate(buckets, start=2):
            ws.cell(row=row, column=1, value=bucket['name'])
            ws.cell(row=row, column=2, value=bucket['created'].strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=3, value=bucket['last_modified'].strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row, column=4, value=bucket['location'])
            ws.cell(row=row, column=5, value=bucket['storage_class'])

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(filename)
        logging.info(f"Results written to {filename}")
    except Exception as e:
        logging.error(f"Error writing to Excel: {str(e)}")

def main():
    args = parse_arguments()
    
    try:
        cutoff_date = datetime.datetime.strptime(args.cutoff_date, "%Y-%m-%d")
    except ValueError:
        logging.error("Invalid date format. Please use YYYY-MM-DD.")
        return

    logging.info(f"Searching for unmodified buckets in project {args.project_id} since {args.cutoff_date}")
    
    try:
        client = get_storage_client(args.project_id)
        unmodified_buckets = get_unmodified_buckets(client, cutoff_date, args.max_workers)
        
        if unmodified_buckets:
            write_to_excel(unmodified_buckets, args.output)
            logging.info(f"Found {len(unmodified_buckets)} unmodified buckets")
        else:
            logging.info("No unmodified buckets found")
    except Exception as e:
        logging.error(f"An error occurred: {str(e)}")

if __name__ == "__main__":
    main()